import requests 
import bs4
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import openpyxl 
import argparse

# Function named crtnewwb- will be called once.
def crtnewwb(f_loc):
# Calling a Workbook() function of openpyx to create a new blank Workbook object and assigning it to variable f_wb 
    f_wb = openpyxl.Workbook()

# Getting workbook's active sheet using the active attribute and assign it to a variable f_sheet
    f_sheet = f_wb.active 

# changing the name of the sheet 
    f_sheet.title = "2nd PUC"

# Adding the Column names
    f_sheet.cell(1,1).value='Reg Num'
    f_sheet.cell(1,2).value='Name'
    f_sheet.cell(1,3).value='Subject-1'
    f_sheet.cell(1,4).value='Subject-2'
    f_sheet.cell(1,5).value='Subject-3'
    f_sheet.cell(1,6).value='Subject-4'
    f_sheet.cell(1,7).value='Subject-5'
    f_sheet.cell(1,8).value='Subject-6'
    f_sheet.cell(1,9).value='Total'
#Saving the file in the location provided by you   
    f_wb.save(f_loc)
    
    # Function to call the workbook and pass in extracted data
def inpt(f_loc,row_num,name,sub1,sub2,sub3,sub4,sub5,sub6,total,reg):
    #opening the workbook
    f_wb = openpyxl.load_workbook(f_loc)
    
    #selecting the sheet where we have to input data
    f_sheet = f_wb['2nd PUC']
    
    #inputing the data inot respective fields
    f_sheet.cell(row_num,1).value = reg
    f_sheet.cell(row_num,2).value = name
    f_sheet.cell(row_num,3).value = sub1
    f_sheet.cell(row_num,4).value = sub2
    f_sheet.cell(row_num,5).value = sub3
    f_sheet.cell(row_num,6).value = sub4
    f_sheet.cell(row_num,7).value = sub5
    f_sheet.cell(row_num,8).value = sub6
    f_sheet.cell(row_num,9).value = total
    
    #keeping a check as to how many rows are entered
    print(row)
    
    #saving the file
    f_wb.save(f_loc)
    
def automation(f_loc,inpt1,inpt2):
    #initializing chrome browser as selenium webdriver to variable browser 
    browser = webdriver.Chrome()
    
    #calling the function to create new workbook
    crtnewwb(f_loc) 
    
    #initializing row =2 as the 1st roe in workbook are the headding
    row = 2
    
    #fetching the HTML of the website
    browser.get('http://karresults.nic.in/indexPUC_2020.asp')
    
    #extracting data for the given range
    for i in range(int(inpt1),(inpt2)):
        try: 
            ele = browser.find_element_by_name('reg')
            ele.clear()
            ele.send_keys(str(i))
            ele.send_keys(Keys.RETURN)
            soup = bs4.BeautifulSoup(browser.page_source,'lxml')
            req_data = soup.select('td')
            name = req_data[1].getText().strip()
            sub1 = req_data[7].getText().strip()
            sub2= req_data[11].getText().strip()
            sub3 =req_data[17].getText().strip()
            sub4 =req_data[21].getText().strip()
            sub5 =req_data[25].getText().strip()
            sub6 =req_data[29].getText().strip()
            gtm =req_data[33].getText().strip()
            inpt(row,name,sub1,sub2,sub3,sub4,sub5,sub6,gtm,i)
            row+=1
            #perform the action of clicking back button
            browser.back()
        except:
            browser.back()
            continue
if __name__ == '__main__':
#initializing variable parser to argument_parser 
    parser = argparse.ArgumentParser() 
#adding 3 arguments: 2 for the range of Registration number and one for file path   
    parser.add_argument("-i1", "--input1", help = "file path of final workbook") 
    parser.add_argument("-i2", "--input2", help = "File to be automated ") 
    parser.add_argument("-fp", "--file_path", help = "File to be automated ") 
    args = parser.parse_args() 
    if args.input1 and args.input2 and args.file_path:
        automation(args.file_path,args.input1,args.input2)
    else:
        print("Required fields not met")